import os
from openpyxl import Workbook, load_workbook


def writing_file_excel(price_list, name_file):
    """ Входные данные - словарь"""
    check_dir(name_file)  # Проверка и создание папки
    wb = Workbook()
    ws = wb.active
    row = 1
    for key, value_list in price_list.items():
        ws.cell(row=row, column=1, value=key)
        column = 2
        for value in value_list:
            ws.cell(row=row, column=column, value=value)
            column += 1
        row += 1

    wb.save(filename=name_file)


def check_dir(p_1):
    head, tail = os.path.split(p_1)
    check = os.path.isdir(head)

    if not check:
        os.mkdir(head)
