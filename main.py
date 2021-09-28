from time import sleep
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import os
import datetime
from openpyxl import Workbook, load_workbook
from settings import *


def web():
    def checking_price(tag):
        try:
            p_1 = i.find_element_by_class_name(tag).text
        except:
            p_1 = 0
        return p_1

    def select_city(city, dv):
        sleep(2)
        dv.find_element_by_class_name('Sh').click()
        sleep(2)
        input_p = dv.find_element_by_class_name('Gx')
        input_p.send_keys(city)
        sleep(2)
        input_select = dv.find_element_by_class_name('Ix')
        sleep(2)
        input_select.click()
        sleep(2)

    dict_product = {}
    head, tail = os.path.split(__file__)
    name_ch = os.path.normpath(f'{head}/chromedriver/chromedriver.exe')
    options = Options()
    options.add_experimental_option('prefs', options_web)
    ch = webdriver.Chrome(name_ch, options=options)

    ch.get("https://www.eldorado.ru/c/smartfony/")
    select_city('Омск', ch)
    sleep(4)
    error_count = 0
    permission = True
    page_number = 1
    while permission:
        try:  # Если каких то данных нет то ждем
            products = ch.find_elements_by_xpath('//*[@id="listing-container"]/ul/li')
            if len(products) == 0:
                permission = False
                break
            for i in products:
                name_product = i.find_element_by_class_name('fE').text
                price_product = checking_price('SP')
                dict_product[name_product] = price_product

            page_number += 1
            ch.get(f"https://www.eldorado.ru/c/smartfony/?page={page_number}")
            sleep(1)
            error_count = 0
        except:
            sleep(2)
            error_count += 1
            print(error_count)
            if error_count > 5:
                ch.get(f"https://www.eldorado.ru/c/smartfony/?page={page_number}")
                sleep(1)
    ch.close()
    return dict_product


def writing_file_excel(price_list, name_f):
    """ Входные данные - словарь"""
    # head, tail = os.path.split(__file__)
    # current_time = datetime.datetime.now().strftime("%d-%m-%y_%H-%M") + '_'
    # name_f = os.path.normpath(f'{head}/data/{name_f}_{current_time}.xlsx')
    # name_f = os.path.normpath(f'{head}/data/{name_f}.xlsx')
    name_f = os.path.normpath(name_f)
    wb = Workbook()
    ws = wb.active
    row = 1
    for key, value_list in price_list.items():
        ws.cell(row=row, column=1, value=key)
        column = 2
        for value in value_list:
            ws.cell(row=row, column=column, value=value)
            column += 1
        row += 1

    wb.save(filename=name_f)


data = web()
writing_file_excel(data, file_path)
